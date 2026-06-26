"""Geracao do relatorio final em .xlsx."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def write_report(rows: list[dict[str, Any]], activity: str, out_dir: str | Path) -> Path:
    """Escreve um xlsx com os resultados.

    `rows` e uma lista de dicts: {idx, id, status, detail}.
    Retorna o caminho do arquivo gerado.
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = out_dir / f"relatorio_elaw_{activity}_{ts}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio"

    headers = ["#", "ID da Pasta", "Status", "Detalhe"]
    ws.append(headers)
    for c, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F1D35")
        cell.alignment = Alignment(horizontal="left")

    for r in rows:
        ws.append([r.get("idx", ""), r.get("id", ""), r.get("status", ""), r.get("detail", "")])

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 70
    ws.freeze_panes = "A2"

    wb.save(out_path)
    return out_path
