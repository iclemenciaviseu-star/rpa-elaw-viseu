"""
Leitura de planilhas (.xlsx, .xls, .csv) com chaves por letra de coluna.

Cada linha vira um dict no formato:
    {"A": "12345", "B": "Joao", "C": "10/05/2026", ...}

A primeira linha e tratada como cabecalho APENAS se a primeira celula
nao for numerica (heuristica simples mas eficaz para listas de IDs).

Implementacao usa openpyxl (xlsx) e csv (csv/tsv) - sem pandas.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


def _col_letters(n: int) -> list[str]:
    """Gera ['A', 'B', ..., 'Z', 'AA', 'AB', ...] de tamanho n."""
    out: list[str] = []
    for i in range(n):
        s = ""
        x = i
        while True:
            s = chr(ord("A") + (x % 26)) + s
            x = x // 26 - 1
            if x < 0:
                break
        out.append(s)
    return out


def _cell_to_str(v) -> str:
    """Converte qualquer valor de celula em string limpa."""
    if v is None:
        return ""
    # datetime
    if hasattr(v, "strftime"):
        return v.strftime("%d/%m/%Y")
    # float que e na verdade inteiro (caso comum em planilhas)
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _read_xlsx(path: Path) -> list[list[str]]:
    """Le um .xlsx e devolve lista de listas (linhas) de strings."""
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([_cell_to_str(c) for c in row])
    wb.close()
    return rows


def _read_csv(path: Path, sep: str) -> list[list[str]]:
    """Le um .csv ou .tsv e devolve lista de listas."""
    rows: list[list[str]] = []
    # Tenta UTF-8 primeiro, depois latin-1
    for enc in ("utf-8-sig", "latin-1"):
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                reader = csv.reader(f, delimiter=sep)
                rows = [[(c or "").strip() for c in row] for row in reader]
            break
        except UnicodeDecodeError:
            continue
    return rows


def read_spreadsheet(path: str | Path) -> list[dict[str, str]]:
    """Le uma planilha e devolve lista de dicts (por letra de coluna)."""
    p = Path(path)
    suffix = p.suffix.lower()

    if suffix in (".xlsx", ".xlsm"):
        raw = _read_xlsx(p)
    elif suffix == ".xls":
        raise ValueError(
            ".xls (Excel antigo) nao suportado. Salve como .xlsx no Excel e tente de novo."
        )
    elif suffix == ".csv":
        raw = _read_csv(p, ",")
    elif suffix == ".tsv":
        raw = _read_csv(p, "\t")
    else:
        raise ValueError(f"Formato nao suportado: {suffix}")

    if not raw:
        return []

    # Detecta cabecalho: primeira celula nao numerica = cabecalho
    first_cell = raw[0][0] if raw[0] else ""
    if first_cell and not first_cell.replace(".", "").replace(",", "").isdigit():
        raw = raw[1:]

    # Determina largura maxima
    max_cols = max((len(r) for r in raw), default=0)
    cols = _col_letters(max_cols)

    rows_out: list[dict[str, str]] = []
    for row in raw:
        # Pula linhas onde a coluna A esta vazia (sem ID)
        if not row or not str(row[0]).strip():
            continue
        d = {cols[i]: str(row[i]).strip() if i < len(row) else "" for i in range(max_cols)}
        rows_out.append(d)
    return rows_out


def extract_ids(rows: Iterable[dict[str, str]]) -> list[str]:
    """Extrai a lista de IDs (coluna A) a partir das linhas."""
    return [r["A"] for r in rows]
